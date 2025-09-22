import os
import boto3
from botocore.exceptions import NoCredentialsError, PartialCredentialsError, ClientError
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from PIL import Image as PILImage
from dotenv import load_dotenv
import io
from collections import Counter
from urllib.parse import urlparse

# Cargar variables de entorno desde .env
load_dotenv()

# --- Configuraciones ---
OUTPUT_FILE = "participaciones_export.xlsx"
IMAGE_WIDTH = 240
IMAGE_HEIGHT = 150

def get_all_items_from_dynamodb(table_name, dynamodb_client):
    items = []
    try:
        paginator = dynamodb_client.get_paginator('scan')
        page_iterator = paginator.paginate(TableName=table_name, PaginationConfig={'PageSize': 100})
        for page in page_iterator:
            items.extend(page['Items'])
        print(f"Se encontraron {len(items)} registros en la tabla '{table_name}'.")
        return items
    except ClientError as e:
        if e.response['Error']['Code'] == 'ResourceNotFoundException':
            print(f"Error Crítico: La tabla '{table_name}' no fue encontrada.")
        else:
            print(f"Error de AWS al escanear la tabla: {e}")
        return None

def deserialize_dynamodb_item(item):
    deserialized = {}
    for key, value in item.items():
        data_type = list(value.keys())[0]
        val = value[data_type]
        if data_type == 'S': deserialized[key] = val
        elif data_type == 'N':
            try: deserialized[key] = int(val)
            except (ValueError, TypeError): deserialized[key] = float(val)
        elif data_type == 'BOOL': deserialized[key] = "Sí" if val else "No"
        elif data_type == 'NULL': deserialized[key] = None
        else: deserialized[key] = str(val)
    return deserialized

def generate_statistics_sheet(wb, sheet_name, items):
    if not items: return
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    title_font, header_font = Font(bold=True, size=14), Font(bold=True)
    current_row = 1
    all_keys = set().union(*(item.keys() for item in items))
    excluded_keys = {'id', 'user_id', 'rut', 'nombre', 'timestamp', 'date', 'time', 'url_img_frontal', 'url_img_trasera', 'is_admin', 'propiedad_id', 'comentario', 'comentarios', 'observaciones', 'unidad'}
    question_keys = sorted([k for k in all_keys if k.lower() not in excluded_keys])
    total_participants = len(items)
    ws.cell(row=current_row, column=1, value="Resumen General de Participación").font = title_font
    current_row += 1
    ws.cell(row=current_row, column=1, value="Total de Participaciones"); ws.cell(row=current_row, column=2, value=total_participants).font = header_font
    current_row += 2
    for key in question_keys:
        ws.cell(row=current_row, column=1, value=f"Estadísticas para: {key.replace('_', ' ').title()}").font = title_font
        current_row += 1
        ws.cell(row=current_row, column=1, value="Respuesta").font = header_font
        ws.cell(row=current_row, column=2, value="Votos").font = header_font
        ws.cell(row=current_row, column=3, value="Porcentaje").font = header_font
        current_row += 1
        counts = Counter(item.get(key, 'N/A') for item in items)
        for option, count in sorted(counts.items()):
            ws.cell(row=current_row, column=1, value=str(option))
            ws.cell(row=current_row, column=2, value=count)
            ws.cell(row=current_row, column=3, value=(count / total_participants)).number_format = '0.00%'
            current_row += 1
        current_row += 1
    print("Hoja de estadísticas generada.")

def process_image_from_stream(image_stream, cell):
    try:
        with PILImage.open(image_stream) as pil_img:
            if pil_img.height > pil_img.width: pil_img = pil_img.rotate(-90, expand=True)
            pil_img.thumbnail((IMAGE_WIDTH, IMAGE_HEIGHT))
            img_byte_arr = io.BytesIO()
            pil_img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            return OpenpyxlImage(img_byte_arr)
    except Exception as e:
        print(f"Error procesando imagen desde stream: {e}")
        return None

def export_to_excel():
    print("Iniciando exportación final...")
    try:
        aws_session = boto3.Session(
            aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
            aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'),
            region_name=os.getenv('AWS_DEFAULT_REGION')
        )
        dynamodb_client = aws_session.client('dynamodb')
        s3_client = aws_session.client('s3')
        table_name = os.getenv('DYNAMODB_TABLE_NAME')
        s3_bucket_name = os.getenv('S3_BUCKET_NAME')
        if not table_name or not s3_bucket_name:
            print("Error Crítico: DYNAMODB_TABLE_NAME o S3_BUCKET_NAME no están en .env")
            return
    except (NoCredentialsError, PartialCredentialsError): print("Error de Autenticación: Credenciales AWS no encontradas."); return

    items_raw = get_all_items_from_dynamodb(table_name, dynamodb_client)
    if not items_raw: print("No se encontraron datos. Finalizando."); return

    processed_items = sorted([deserialize_dynamodb_item(item) for item in items_raw], key=lambda x: str(x.get('unidad', 'zzzzzzzz')).lower())
    print(f"{len(processed_items)} filas procesadas y ordenadas por 'unidad'.")

    all_keys = set().union(*(item.keys() for item in processed_items))
    priority_headers = ['id', 'rut', 'nombre', 'unidad']
    image_headers, url_headers = ["Imagen Frontal", "Imagen Trasera"], ["url_img_frontal", "url_img_trasera"]
    other_headers = sorted([k for k in all_keys if k not in set(priority_headers + url_headers)])
    final_headers = priority_headers + other_headers + url_headers + image_headers

    wb = Workbook(); ws = wb.active; ws.title = "Participaciones"; ws.append(final_headers)
    header_to_col_idx = {h: i for i, h in enumerate(final_headers, 1)}
    for col_idx, h in enumerate(final_headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = (IMAGE_WIDTH / 7) if h in image_headers else max(len(h) + 2, 25)
    
    static_folder_path = os.path.join(os.getcwd(), 'static')

    for row_idx, item in enumerate(processed_items, start=2):
        ws.row_dimensions[row_idx].height = IMAGE_HEIGHT * 0.75
        for header, col_idx in header_to_col_idx.items(): ws.cell(row=row_idx, column=col_idx, value=item.get(header, "N/A"))

        image_locations = {"Imagen Frontal": item.get("url_img_frontal"), "Imagen Trasera": item.get("url_img_trasera")}

        for img_header, img_url in image_locations.items():
            col_num = header_to_col_idx.get(img_header)
            if not col_num: continue
            cell = ws.cell(row=row_idx, column=col_num)
            if not img_url or img_url == "N/A": cell.value = "URL no disponible"; continue

            img_to_add = None
            if img_url.startswith('http'): # Es una URL de S3
                try:
                    parsed_url = urlparse(img_url)
                    object_key = parsed_url.path.lstrip('/')
                    response = s3_client.get_object(Bucket=s3_bucket_name, Key=object_key)
                    img_to_add = process_image_from_stream(io.BytesIO(response['Body'].read()), cell)
                    if not img_to_add: cell.value = "Error procesando S3"
                except ClientError as e:
                    if e.response['Error']['Code'] == 'NoSuchKey': cell.value = "No en S3"
                    else: cell.value = f"Error S3: {e.response['Error']['Code']}"; print(f"Error S3: {e}")
            else: # Es una ruta local
                local_path = os.path.join(static_folder_path, img_url.lstrip('/'))
                if os.path.exists(local_path):
                    img_to_add = process_image_from_stream(local_path, cell)
                    if not img_to_add: cell.value = "Error procesando local"
                else: cell.value = "No encontrada local"

            if img_to_add: ws.add_image(img_to_add, cell.coordinate)

    generate_statistics_sheet(wb, "Estadísticas", processed_items)
    try:
        wb.save(OUTPUT_FILE)
        print(f"\n¡Éxito! Exportación finalizada en '{OUTPUT_FILE}'.")
    except Exception as e: print(f"Error al guardar el archivo: {e}")

if __name__ == "__main__":
    export_to_excel()